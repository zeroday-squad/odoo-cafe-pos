import io
from django.db.models import Sum, Count, Avg
from django.db.models.functions import TruncDate
from django.http import FileResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from core.models import Order, OrderItem
from core.permissions import IsAdminUser
from core.utils import apply_report_filters

class SummaryView(APIView):
    """GET /api/reports/summary/"""
    permission_classes = [IsAdminUser]

    def get(self, request):
        qs  = apply_report_filters(Order.objects.filter(status='paid'),
                                    request.query_params)
        agg = qs.aggregate(
            total_orders    = Count('id'),
            revenue         = Sum('total'),
            avg_order_value = Avg('total'),
        )
        kitchen_queue = Order.objects.filter(
            kds_status__in=['to_cook', 'preparing']
        ).count()
        return Response({
            'total_orders':    agg['total_orders']    or 0,
            'revenue':         str(agg['revenue']     or 0),
            'avg_order_value': str(agg['avg_order_value'] or 0),
            'kitchen_queue':   kitchen_queue,
        })


class SalesTrendView(APIView):
    """GET /api/reports/sales-trend/"""
    permission_classes = [IsAdminUser]

    def get(self, request):
        qs = apply_report_filters(Order.objects.filter(status='paid'),
                                   request.query_params)
        data = (
            qs.annotate(date=TruncDate('created_at'))
              .values('date')
              .annotate(revenue=Sum('total'), order_count=Count('id'))
              .order_by('date')
        )
        return Response([
            {
                'date':        str(row['date']),
                'revenue':     str(row['revenue']),
                'order_count': row['order_count'],
            }
            for row in data
        ])


class TopProductsView(APIView):
    """GET /api/reports/top-products/"""
    permission_classes = [IsAdminUser]

    def get(self, request):
        qs = apply_report_filters(Order.objects.filter(status='paid'),
                                   request.query_params)
        data = (
            OrderItem.objects.filter(order__in=qs)
                .values('product__name')
                .annotate(quantity_sold=Sum('quantity'), revenue=Sum('line_total'))
                .order_by('-revenue')[:10]
        )
        return Response([
            {
                'name':          row['product__name'],
                'quantity_sold': row['quantity_sold'],
                'revenue':       str(row['revenue']),
            }
            for row in data
        ])


class TopCategoriesView(APIView):
    """GET /api/reports/top-categories/"""
    permission_classes = [IsAdminUser]

    def get(self, request):
        qs = apply_report_filters(Order.objects.filter(status='paid'),
                                   request.query_params)
        data = (
            OrderItem.objects.filter(order__in=qs)
                .values('product__category__name')
                .annotate(revenue=Sum('line_total'))
                .order_by('-revenue')[:10]
        )
        return Response([
            {
                'name':    row['product__category__name'],
                'revenue': str(row['revenue']),
            }
            for row in data
        ])


class TopOrdersView(APIView):
    """GET /api/reports/top-orders/"""
    permission_classes = [IsAdminUser]

    def get(self, request):
        qs = apply_report_filters(Order.objects.filter(status='paid'),
                                   request.query_params)
        data = (
            qs.order_by('-total')
              .values('order_number', 'customer__name', 'total')[:10]
        )
        return Response([
            {
                'order_number':  row['order_number'],
                'customer_name': row['customer__name'] or 'Walk-in',
                'total':         str(row['total']),
            }
            for row in data
        ])


class ExportView(APIView):
    """
    GET /api/reports/export/?format=pdf|xlsx
    Returns a downloadable file. Default format is pdf.
    """
    permission_classes = [IsAdminUser]

    def get(self, request):
        fmt = request.query_params.get('format', 'pdf')
        qs  = apply_report_filters(Order.objects.filter(status='paid').order_by('-created_at'),
                                    request.query_params)
        orders = list(qs.values('order_number', 'total', 'created_at', 'status'))

        if fmt == 'xlsx':
            return self._export_xlsx(orders)
        return self._export_pdf(orders)

    def _export_xlsx(self, orders):
        import openpyxl
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # Headers
        headers = ["Order Number", "Total", "Created At", "Status"]
        ws.append(headers)

        # Data rows
        for o in orders:
            dt_str = o['created_at'].strftime('%Y-%m-%d %H:%M:%S') if o['created_at'] else ''
            ws.append([
                o['order_number'],
                float(o['total']),
                dt_str,
                o['status']
            ])

        # Auto-adjust column dimensions
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = FileResponse(
            buffer,
            as_attachment=True,
            filename="sales_report.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        return response

    def _export_pdf(self, orders):
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=18,
            leading=22,
            spaceAfter=12
        )

        elements.append(Paragraph("CafeFlow POS - Sales Report", title_style))
        elements.append(Spacer(1, 10))

        # Build table data
        data = [["Order Number", "Total", "Created At", "Status"]]
        for o in orders:
            dt_str = o['created_at'].strftime('%Y-%m-%d %H:%M:%S') if o['created_at'] else ''
            data.append([
                o['order_number'],
                f"${o['total']}",
                dt_str,
                o['status']
            ])

        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(t)
        doc.build(elements)

        buffer.seek(0)
        response = FileResponse(
            buffer,
            as_attachment=True,
            filename="sales_report.pdf",
            content_type="application/pdf"
        )
        return response
